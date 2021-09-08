import shutil
import openpyxl
import warnings
import requests
import os
import re
import time
import glob
from tkinter import *
from selenium import webdriver



def retrive_number(IP):
    l = []
    url = "http://atllchickentools.scientificgames.com/memprofileserver/"
    tes = requests.get(url).text
    with open('line.txt', 'w') as f:
        f.write(tes)
    p = open('line.txt')
    for position, line in enumerate(p):
        if IP in line:
            op = line.split("?")
            io = op[len(op) - 1]
            ri = io.split("_")
            break

    return ri[0]
def chromedriver_download_path_change(download_path,driver_path):
    chromeoptions = webdriver.ChromeOptions()
    prefs = {"download.default_directory": download_path}
    chromeoptions.add_experimental_option("prefs", prefs)
    return webdriver.Chrome(executable_path=driver_path, options=chromeoptions)
def latest_download_file():
    path = r'C:\memory'
    os.chdir(path)
    files = sorted(os.listdir(os.getcwd()), key=os.path.getmtime)
    newest = files[-1]

    return newest
def waiting_time_for_download():
    fileends = "crdownload"
    while "crdownload" == fileends:
        time.sleep(1)
        newest_file = latest_download_file()
        if "crdownload" in newest_file:
            fileends = "crdownload"
        else:
            fileends = "none"
def getting_lattest_file():
    list_of_files = glob.glob('C:\memory\*')  # * means all if need specific format then *.csv
    latest_file = max(list_of_files, key=os.path.getctime)
    return latest_file
def memory_extraction(path):
    p = []
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    maxrow = sheet_obj.max_row
    for i in range(11, maxrow):
        cell_obj = sheet_obj.cell(row=i, column=4)
        p.append(cell_obj.value)
    p.sort()
    no_of_d = (len(p) / 12)
    days = int(no_of_d) - 1
    final = (p[len(p) - 1] - p[0]) / days
    ou = final / 1024
    print("Memory leak for {0}days ={1}Mb".format(days, float(ou)))
    Ro="Memory leak for {0}days ={1}Mb".format(days, float(ou))
    s=Label(master,text=Ro)
    s.grid(row=7,column=0,sticky=W)
    warnings.warn("deprecated", DeprecationWarning)
def memory_leak(path):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        memory_extraction(path)
def leak():
    path = r'C:\Users\anurag.phad\PycharmProjects\CHromewebdrivechickentool\venv\Scripts\chromedriver.exe'
    download_path = "C:\memory"
    IP = input_ip.get()
    Unique_number = retrive_number(IP)
    driver = chromedriver_download_path_change(download_path, path)
    driver.get("http://atllchickentools.scientificgames.com/memprofileserver/makexls_argos.cgi?dat={0}_{1}".format(Unique_number, IP))
    waiting_time_for_download()
    driver.close()
    memory_leak(getting_lattest_file())







def Creation_of_folders():
    if os.path.exists("C:\\EGM"):
        print("")
    else:
        os.mkdir("C:\\EGM")
        print("EGM Folder Get created")

    if os.path.exists("C:\\memory"):
        print("")
    else:
        os.mkdir("C:\\memory")
        print("MemoryFolder Get created")


    if os.path.exists("C:\\EGM\\Game_logs"):
        print(" ")
    else:
        os.mkdir("C:\\EGM\\Game_logs")

    if os.path.exists("C:\\EGM\\Full gaffer"):
        print("")
    else:
        os.mkdir("C:\\EGM\\Full gaffer")
def Game_name(ip):
    l=[]
    p=[]
    url="http://{0}/Logs/Game/".format(ip)
    page=requests.get(url).text
    res = re.findall(r'\w+',page)
    p=res
    return p[39]
def Ip_access(input):

    #-------------Game logs Files-------------------------#
    server_url = "http://{0}/Logs/Game/{1}/Logs/{2}_Server.log".format(input,Game_name(input),Game_name(input))
    client_url = "http://{0}/Logs/Game/{1}/Logs/{2}_clientlog.log".format(input,Game_name(input),Game_name(input))
    profil_url = "http://{0}/Logs/Game/{1}/Logs/{2}_Profiling.log".format(input,Game_name(input),Game_name(input))
    client_page=requests.get(client_url).text
    server_page = requests.get(server_url).text
    profile_page = requests.get(profil_url).text
    with open('C:\EGM\Game_logs\{0}_Server.txt'.format(Game_name(input)), 'w') as f:
        f.write(server_page)
    with open('C:\EGM\Game_logs\{0}_Client.txt'.format(Game_name(input)), 'w') as f:
        f.write(client_page)
    with open('C:\EGM\Game_logs\{0}_Profile.txt'.format(Game_name(input)), 'w') as f:
        f.write(profile_page)
    #--------------Game full gaffer---------------------#
    full_gaffer_url = "http://{0}/Logs/Game/{1}/FullGaffer/{2}_LastPlayedGames.xml".format(input,Game_name(input),Game_name(input))
    full_gaffer_page=requests.get(full_gaffer_url).text
    with open('C:\EGM\Full gaffer\{0}_Last_played_game.xml'.format(Game_name(input)),'w') as f:
        f.write(full_gaffer_page)

    # #-----------------Health Monitor-------------------#
    # Health = "http://{0}/Logs/HealthMonitor.log".format(input)
    # Health_page = requests.get(Health).text
    # with open('C:\EGM\HealthMonitor.log', 'w') as f:
    #     f.write(Health_page)

    #-----------------Process Manager-------------------#
    Process = "http://{0}/Logs/ProcessManager.log".format(input)
    Process_page = requests.get(Process).text
    with open('C:\EGM\ProcessManager.log.', 'w') as f:
        f.write(Process_page)

def save():
    # final=Label(master,text="Processing.......")
    # final.grid(row=4,column=0,sticky=W)
    print(input_ip.get())
    print("Procesing.......")
    Ip_access(input_ip.get())
    shutil.make_archive("{0}-{1}".format(Game_name(input_ip.get()), int(time.time())), "zip", root_dir='c:/EGM')
    time.sleep(5)
    shutil.rmtree("C:\\EGM\\Game_logs\\")
    shutil.rmtree("C:\\EGM\\Full gaffer\\")
    print("Your file is ready")
    save="EGM logs zip file\n{0}".format(os.getcwd())
    Label(master,text=save).grid(row=8,column=0,sticky=W)


def serach():
    print("""    Script has started and in process.......
    You can minimise this tab and it will continously tracking the EGM for any crash""")
    url = "http://{0}/Logs/HealthMonitor.log".format(input_ip.get())
    word = "HealthMonitorSvc::ShowFaultScreen()"
    page = requests.get(url).text
    while True:
        if word in page:
            break
    print("Game has been crashed please check the EGM of IP address-{0}".format(input_ip.get()))
    t="Game has been crashed please check the EGM of IP address-{0}".format(input_ip.get())
    s=Label(master,text=t)
    s.grid(row=10,column=0,sticky=W)




if __name__=="__main__":
    master = Tk()
    master.title("EGM Tool")
    master.geometry('900x400')
    Creation_of_folders()
    Titile=Label(master,text="Autocasino Tool")
    Titile.grid(row=0,column=0,padx=120)
    EGMip = Label(master, text="Enter EGM IP")
    EGMip.grid(row=1, column=0,sticky=W)
    input_ip=Entry(master)
    input_ip.grid(row=2,column=0,sticky=W)
    B=Button(master,text="Memory Leak",command=leak)
    B.grid(row=3,column=0,padx=10)
    Info = Label(master, text="""! Directly memory leak of the EGM can be found
      """)
    Info.grid(row=3, column=1, padx=10)
    P=Button(master,text="Create log file",command=save)
    P.grid(row=4, column=0, padx=10)
    Info = Label(master, text="""! EGM logs get created in ZIP
     """)
    Info.grid(row=4, column=1,padx=10)
    R = Button(master, text="Start Script AC Crash",command=serach)
    R.grid(row=5, column=0, padx=10)
    Info = Label(master, text="""! AC script for crash detection
     Note:-Do Not close this tab
     Note:-It will RUN continously and when crash observed error will appear on Screen""")
    Info.grid(row=5, column=1,sticky=W)



    master.mainloop()











