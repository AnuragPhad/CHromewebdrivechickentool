import openpyxl
import warnings



def memory_extraction(path):
    p = []
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    maxrow = sheet_obj.max_row
    for i in range(11, maxrow):
        cell_obj = sheet_obj.cell(row=i, column=4)
        p.append(cell_obj.value)
    p.sort()
    print(p)
    no_of_d = (len(p) / 12)
    days = int(no_of_d) - 1
    final = (p[len(p) - 1] - p[0]) / days
    ou = final / 1024
    print("Memory leak for {0}days ={1}Mb".format(days, float(ou)))
    warnings.warn("deprecated", DeprecationWarning)

def memory_leak(path):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        memory_extraction(path)

memory_leak('jk.xlsm')














