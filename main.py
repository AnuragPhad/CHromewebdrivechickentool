import openpyxl
import warnings
import requests
import os
import time
import glob
from tkinter import *
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg,NavigationToolbar2Tk)
from selenium import webdriver

IP=input("Enter the IP address")
#function to retrive the unique number from chicken tool
def retrive_number(IP):
    l = []
    url = "http://atllchickentools.scientificgames.com/memprofileserver/"
    tes = requests.get(url).text
    with open('line.txt', 'w') as f:
        f.write(tes)
    p = open('line.txt')
    for position, line in enumerate(p):
        if IP in line:
            op = line.split("?")
            io = op[len(op) - 1]
            ri = io.split("_")
            break

    return ri[0]


Unique_number=retrive_number(IP)

#Path for the chrome driver and download location
path=r'C:\Users\anurag.phad\PycharmProjects\CHromewebdrivechickentool\venv\Scripts\chromedriver.exe'
download_path="C:\memory"


#function to create the download location

def chromedriver_download_path_change(download_path,driver_path):
    chromeoptions = webdriver.ChromeOptions()
    prefs = {"download.default_directory": download_path}
    chromeoptions.add_experimental_option("prefs", prefs)
    return webdriver.Chrome(executable_path=driver_path, options=chromeoptions)

driver=chromedriver_download_path_change(download_path,path)
driver.get("http://atllchickentools.scientificgames.com/memprofileserver/makexls_argos.cgi?dat={0}_{1}".format(Unique_number,IP))


def latest_download_file():
    path = r'C:\memory'
    os.chdir(path)
    files = sorted(os.listdir(os.getcwd()), key=os.path.getmtime)
    newest = files[-1]

    return newest


def waiting_time_for_download():
    fileends = "crdownload"
    while "crdownload" == fileends:
        time.sleep(1)
        newest_file = latest_download_file()
        if "crdownload" in newest_file:
            fileends = "crdownload"
        else:
            fileends = "none"


waiting_time_for_download()
driver.close()

def getting_lattest_file():
    list_of_files = glob.glob('C:\memory\*')  # * means all if need specific format then *.csv
    latest_file = max(list_of_files, key=os.path.getctime)
    return latest_file


def memory_extraction(path):
    p = []
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    maxrow = sheet_obj.max_row
    for i in range(11, maxrow):
        cell_obj = sheet_obj.cell(row=i, column=4)
        p.append(cell_obj.value)
    p.sort()
    no_of_d = (len(p) / 12)
    days = int(no_of_d) - 1
    final = (p[len(p) - 1] - p[0]) / days
    ou = final / 1024
    print("Memory leak for {0}days ={1}Mb".format(days, float(ou)))
    warnings.warn("deprecated", DeprecationWarning)


def memory_leak(path):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        memory_extraction(path)


memory_leak(getting_lattest_file())





